"""Transaction source connectors: CSV/XLSX files and REST APIs.

Lives in shared (not the worker) because the API's import wizard needs the
same parser for synchronous dry-run previews. Both connectors produce the
same canonical row shape; config schemas are documented in mocks/README.md
and seeded from mocks/data/mapping_template.json.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any
from zoneinfo import ZoneInfo

_ENCODINGS = {
    "utf-8": "utf-8",
    "utf-8-sig": "utf-8-sig",
    "big5": "big5hkscs",
    "gb18030": "gb18030",
}
_DEFAULT_DATE_FORMAT = "%Y-%m-%d %H:%M:%S"
# Trailing timezone abbreviation some exports append, e.g. "2025-11-18 09:40:32 HKT".
_TZ_SUFFIX = re.compile(r"\s+([A-Za-z]{2,5})$")
_HK_TZ = ZoneInfo("Asia/Hong_Kong")
_ET_TZ = ZoneInfo("America/New_York")
# Fallback datetime/date formats tried in addition to the config date_format, so a
# range of real broker exports parse without per-file tuning.
_DT_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y/%m/%d %H:%M:%S",
    "%d/%m/%Y %H:%M:%S",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
)
_DATE_FORMATS = ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y")

CANONICAL_KEYS = (
    "trade_date",
    "ext_txn_id",
    "ordered_at",
    "executed_at",
    "broker_code",
    "client_account",
    "client_name",
    "stock_code",
    "stock_name",
    "side",
    "quantity",
    "price",
    "amount",
    "channel",
    "status",
)


@dataclass
class CanonicalTxn:
    ext_txn_id: str | None = None
    trade_date: date | None = None  # per-row date when the file carries a date column
    ordered_at: datetime | None = None
    executed_at: datetime | None = None
    broker_code: str | None = None
    client_account: str | None = None
    client_name: str | None = None
    stock_code: str | None = None
    stock_name: str | None = None
    side: str | None = None  # buy | sell
    quantity: float | None = None
    price: float | None = None
    amount: float | None = None
    channel: str | None = None  # phone | online | None
    raw: dict[str, Any] = field(default_factory=dict)
    skip_reason: str | None = None  # status | side | blank | duplicate


def _norm_stock_code(value: str | None) -> str | None:
    if not value:
        return None
    text = str(value).strip().upper()
    if any("A" <= ch <= "Z" for ch in text):
        # US securities use alphabetic tickers (NVDA, BRK.B, BF-B). Keeping
        # only digits silently erased every US order from reconciliation.
        ticker = re.sub(r"[^A-Z0-9.-]", "", text)
        return ticker or None
    digits = "".join(ch for ch in text if ch.isdigit()).lstrip("0")
    return digits or None


def _to_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    while len(text) >= 2 and text[0] == text[-1] and text[0] in {"'", '"'}:
        text = text[1:-1].strip()
    return text or None


def _split_tz_suffix(value: str) -> tuple[str, str | None]:
    text = value.strip()
    match = _TZ_SUFFIX.search(text)
    if not match:
        return text, None
    return text[: match.start()].strip(), match.group(1).upper()


def _source_tz(suffix: str | None, default: ZoneInfo) -> ZoneInfo:
    if suffix == "HKT":
        return _HK_TZ
    if suffix in {"ET", "EST", "EDT"}:
        return _ET_TZ
    return default


def _parse_dt(value: Any, fmt: str, tz: ZoneInfo) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):  # openpyxl gives real datetimes
        parsed = value.replace(tzinfo=tz) if value.tzinfo is None else value
        return parsed.astimezone(tz)
    if isinstance(value, date):  # date-only cell
        return datetime(value.year, value.month, value.day, tzinfo=tz)
    text, suffix = _split_tz_suffix(str(value))
    source_tz = _source_tz(suffix, tz)
    for f in (fmt, *(x for x in _DT_FORMATS if x != fmt)):
        try:
            return datetime.strptime(text, f).replace(tzinfo=source_tz).astimezone(tz)
        except ValueError:
            continue
    return None


def _parse_date(value: Any) -> date | None:
    """Parse a date-only column value (the per-row trade date)."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text, _suffix = _split_tz_suffix(str(value))
    for f in (*_DATE_FORMATS, "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, f).date()
        except ValueError:
            continue
    return None


def _map_value(value: Any, mapping: dict[str, list[str]] | None) -> str | None:
    """'B' -> 'buy' given {'buy': ['B', 'BUY'], ...}. None when unmapped."""
    if value is None or not mapping:
        return None
    needle = str(value).strip().upper()
    for canonical, variants in mapping.items():
        if needle in {str(v).strip().upper() for v in variants}:
            return canonical
    return None


def canonicalize(values: dict[str, Any], config: dict[str, Any]) -> CanonicalTxn:
    """Turn one extracted row (canonical keys -> raw values) into a CanonicalTxn."""
    tz = ZoneInfo(config.get("timezone", "Asia/Hong_Kong"))
    fmt = config.get("date_format", _DEFAULT_DATE_FORMAT)

    txn = CanonicalTxn(raw={k: v for k, v in values.items() if v not in (None, "")})

    if not txn.raw:  # fully blank row (e.g. spacer rows between orders)
        txn.skip_reason = "blank"
        return txn

    status_filter = config.get("status_filter") or {}
    include = status_filter.get("include")
    if include:
        status_value = str(values.get("status") or "").strip().upper()
        if status_value not in {str(s).upper() for s in include}:
            txn.skip_reason = "status"
            return txn

    side = _map_value(values.get("side"), config.get("side_values"))
    if side not in ("buy", "sell"):
        txn.skip_reason = "side"
        return txn
    txn.side = side

    txn.ext_txn_id = _clean_text(values.get("ext_txn_id"))
    txn.trade_date = _parse_date(values.get("trade_date"))
    txn.ordered_at = _parse_dt(values.get("ordered_at"), fmt, tz)
    txn.executed_at = _parse_dt(values.get("executed_at"), fmt, tz)
    txn.broker_code = _clean_text(values.get("broker_code"))
    txn.client_account = _clean_text(values.get("client_account"))
    txn.client_name = _clean_text(values.get("client_name"))
    txn.stock_code = _norm_stock_code(values.get("stock_code"))
    txn.stock_name = _clean_text(values.get("stock_name"))
    txn.quantity = _to_number(values.get("quantity"))
    txn.price = _to_number(values.get("price"))
    txn.amount = _to_number(values.get("amount"))
    txn.channel = _map_value(values.get("channel"), config.get("channel_values"))
    return txn


def _extract(row: dict[str, Any], column_mapping: dict[str, str]) -> dict[str, Any]:
    values = {key: row.get(column) for key, column in column_mapping.items()}
    # status may be mapped via status_filter.column instead of column_mapping
    return values


# ---------------------------------------------------------------------------
# CSV / XLSX
# ---------------------------------------------------------------------------


def parse_file(filename: str, data: bytes, config: dict[str, Any]) -> list[CanonicalTxn]:
    if filename.lower().endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(data, config)
    return _parse_csv(data, config)


def _rows_to_txns(rows: list[dict[str, Any]], config: dict[str, Any]) -> list[CanonicalTxn]:
    column_mapping = config["column_mapping"]
    status_column = (config.get("status_filter") or {}).get("column")
    # Some exports emit one row per partial fill of the same order; collapse them
    # to the first row per external id so one order becomes one transaction.
    collapse = bool(config.get("collapse_by_ext_id"))
    suffix_duplicate_ext_id = bool(config.get("suffix_duplicate_ext_id"))
    out: list[CanonicalTxn] = []
    seen_refs: dict[str, int] = {}
    for row in rows:
        values = _extract(row, column_mapping)
        if status_column:
            values["status"] = row.get(status_column)
        txn = canonicalize(values, config)
        if collapse and txn.skip_reason is None and txn.ext_txn_id:
            if txn.ext_txn_id in seen_refs:
                txn.skip_reason = "duplicate"
            else:
                seen_refs[txn.ext_txn_id] = 1
        elif suffix_duplicate_ext_id and txn.skip_reason is None and txn.ext_txn_id:
            seen_refs[txn.ext_txn_id] = seen_refs.get(txn.ext_txn_id, 0) + 1
            if seen_refs[txn.ext_txn_id] > 1:
                txn.ext_txn_id = f"{txn.ext_txn_id}-{seen_refs[txn.ext_txn_id]}"
        out.append(txn)
    return out


def detected_trade_dates(txns: list[CanonicalTxn]) -> list[date]:
    """Distinct per-row trade dates among importable rows (empty if none mapped)."""
    return sorted({t.trade_date for t in txns if t.skip_reason is None and t.trade_date})


def _parse_csv(data: bytes, config: dict[str, Any]) -> list[CanonicalTxn]:
    configured = _ENCODINGS.get(config.get("encoding", "utf-8"), "utf-8")
    encodings = [configured, *(e for e in ("utf-8-sig", "gb18030", "big5hkscs") if e != configured)]
    header_row = int(config.get("header_row", 1))
    expected_headers = set(config["column_mapping"].values())
    best: tuple[int, str] | None = None
    for encoding in encodings:
        try:
            candidate = data.decode(encoding)
        except UnicodeDecodeError:
            continue
        lines = candidate.splitlines()
        if len(lines) < header_row:
            continue
        headers = set(next(csv.reader([lines[header_row - 1]]), []))
        score = len(headers & expected_headers)
        if best is None or score > best[0]:
            best = (score, candidate)
        if score == len(expected_headers):
            break
    if best is None:
        raise UnicodeDecodeError(configured, data, 0, 1, "unsupported CSV encoding")
    text = best[1]
    lines = text.splitlines()
    reader = csv.DictReader(io.StringIO("\n".join(lines[header_row - 1 :])))
    return _rows_to_txns(list(reader), config)


def _parse_xlsx(data: bytes, config: dict[str, Any]) -> list[CanonicalTxn]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = config.get("sheet_name")
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    for _ in range(int(config.get("header_row", 1)) - 1):  # skip any preamble rows
        next(rows_iter, None)
    header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    rows = [dict(zip(header, r, strict=False)) for r in rows_iter]
    wb.close()
    return _rows_to_txns(rows, config)


# ---------------------------------------------------------------------------
# REST API
# ---------------------------------------------------------------------------

MAX_PAGES = 100


def fetch_api(
    config: dict[str, Any], credential: str | None, trade_date: date
) -> list[CanonicalTxn]:
    """Pull one trade date from a REST source. Raises on transport errors."""
    import httpx

    base_url = config["base_url"].rstrip("/")
    path = config["path_template"].format(date=trade_date.isoformat())
    pagination = config.get("pagination") or {}
    page_param = pagination.get("page_param", "page")
    size_param = pagination.get("size_param", "page_size")
    page_size = int(pagination.get("page_size", 100))
    items_field = pagination.get("items_field", "items")
    total_field = pagination.get("total_field")

    headers: dict[str, str] = {}
    auth = None
    kind = config.get("auth_kind", "none")
    if kind == "api_key_header" and credential:
        headers[config.get("auth_header", "X-API-Key")] = credential
    elif kind == "bearer" and credential:
        headers["Authorization"] = f"Bearer {credential}"
    elif kind == "basic" and credential:
        user, _, password = credential.partition(":")
        auth = (user, password)

    field_mapping = config["field_mapping"]
    sep = "&" if "?" in path else "?"
    txns: list[CanonicalTxn] = []
    with httpx.Client(timeout=float(config.get("timeout_s", 30)), auth=auth) as client:
        for page in range(1, MAX_PAGES + 1):
            url = f"{base_url}{path}{sep}{page_param}={page}&{size_param}={page_size}"
            resp = client.get(url, headers=headers)
            resp.raise_for_status()
            body = resp.json()
            items = body.get(items_field) or []
            for item in items:
                values = {key: item.get(json_key) for key, json_key in field_mapping.items()}
                txns.append(canonicalize(values, config))
            total = body.get(total_field) if total_field else None
            if not items or (total is not None and page * page_size >= int(total)):
                break
    return txns
